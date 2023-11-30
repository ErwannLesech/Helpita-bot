from dotenv import load_dotenv
import os
from openpyxl import load_workbook
import time
import pandas as pd 

load_dotenv(dotenv_path="config.conf")

log_file_path = os.getenv("LOG_FILE_PATH")
price_file_path = os.getenv("PRICE_FILE_PATH")

def count_words(message):
    return len(message.split(" "))

def create_log(message, user):
    init_logger()
    price = ((count_words(message) / 4) * 3) / 1000
    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
    write_in_excel(user, current_time.split(" ")[0], current_time.split(" ")[1], message, price)
    wb.save(log_file_path)
    update_price(user, price)

# Excel logging

wb = load_workbook(log_file_path)
ws = wb.active

wb2 = load_workbook(price_file_path)
ws2 = wb2.active

def write_in_excel(user, date, time, message, price):
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value=(str)(user))
    ws.cell(row=last_row, column=2, value=date)
    ws.cell(row=last_row, column=3, value=time)
    ws.cell(row=last_row, column=4, value=message)
    ws.cell(row=last_row, column=5, value=price)


def init_logger():
    ws.cell(row=1, column=1, value="user")
    ws.cell(row=1, column=2, value="date")
    ws.cell(row=1, column=3, value="time")
    ws.cell(row=1, column=4, value="message")
    ws.cell(row=1, column=5, value="price")

def update_price(user, price):    
    ws2.cell(row=1, column=1, value="user name")
    ws2.cell(row=1, column=2, value="price")
    
    count = 0
    for row in ws2.iter_rows():
        if count > 5:
            break
        count += 1
        if row[0].value == (str)(user):
            value = (float)(row[1].value) + price
            row[1].value = (str)(value)
            wb2.save(price_file_path)
            break

    if count > 5:
        for row in ws2.iter_rows():
            if row[0].value == None:
                row[0].value = (str)(user);
                row[1].value = (str)(price)
                break
    
    wb2.save(price_file_path)